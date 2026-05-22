from __future__ import annotations

import io
from typing import Any

import pandas as pd


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def build_excel_workbook(
    summary: pd.DataFrame,
    matched_transactions: pd.DataFrame,
    raw_matched_rows: pd.DataFrame,
    campaign_mapping: pd.DataFrame,
    ttd_unmatched: pd.DataFrame,
    ga4_unmatched: pd.DataFrame,
    data_quality: pd.DataFrame,
    extra_sheets: dict[str, pd.DataFrame] | None = None,
) -> bytes:
    output = io.BytesIO()
    sheets: dict[str, pd.DataFrame] = {
        "Summary": summary,
        "Matched Transactions": matched_transactions,
        "Raw Matched Rows": raw_matched_rows,
        "Campaign Mapping": campaign_mapping,
        "TTD Unmatched": ttd_unmatched,
        "GA4 Unmatched": ga4_unmatched,
        "Data Quality": data_quality,
    }
    if extra_sheets:
        sheets.update(extra_sheets)

    engine = "xlsxwriter" if _has_xlsxwriter() else "openpyxl"
    writer_kwargs = {"engine": engine}
    if engine == "xlsxwriter":
        writer_kwargs["datetime_format"] = "yyyy-mm-dd hh:mm:ss"

    with pd.ExcelWriter(output, **writer_kwargs) as writer:
        for sheet_name, df in sheets.items():
            safe_name = _safe_sheet_name(sheet_name)
            export_df = _coerce_excel_safe(df)
            export_df.to_excel(writer, sheet_name=safe_name, index=False)
            if engine == "xlsxwriter":
                _format_xlsxwriter_sheet(writer, safe_name, export_df)
            else:
                _format_openpyxl_sheet(writer, safe_name, export_df)

    return output.getvalue()


def _has_xlsxwriter() -> bool:
    try:
        import xlsxwriter  # noqa: F401
    except ImportError:
        return False
    return True


def _coerce_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    output = df.copy()
    for column in output.columns:
        if pd.api.types.is_datetime64_any_dtype(output[column]):
            output[column] = output[column].dt.tz_localize(None)
    return output


def _format_xlsxwriter_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    header_format = workbook.add_format({"bold": True, "bg_color": "#EAECEF", "border": 1})
    currency_format = workbook.add_format({"num_format": "$#,##0.00"})
    percentage_format = workbook.add_format({"num_format": "0.0%"})
    integer_format = workbook.add_format({"num_format": "#,##0"})

    worksheet.freeze_panes(1, 0)
    if len(df.columns) > 0:
        worksheet.autofilter(0, 0, max(len(df), 1), len(df.columns) - 1)

    for col_idx, column in enumerate(df.columns):
        worksheet.write(0, col_idx, column, header_format)
        width = _column_width(df[column], column)
        fmt: Any = None
        lower = str(column).lower()
        if "revenue" in lower or "monetary value" in lower:
            fmt = currency_format
        elif "share" in lower or "rate" in lower:
            fmt = percentage_format
        elif column in {"Conversions", "Raw TTD rows", "Unique transaction IDs", "Matched transactions"}:
            fmt = integer_format
        worksheet.set_column(col_idx, col_idx, width, fmt)


def _format_openpyxl_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    from openpyxl.styles import Border, Font, PatternFill, Side

    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"
    if len(df.columns) > 0:
        worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="EAECEF")
    thin = Side(style="thin", color="D0D7DE")
    header_border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = header_border

    for col_idx, column in enumerate(df.columns, start=1):
        letter = worksheet.cell(row=1, column=col_idx).column_letter
        worksheet.column_dimensions[letter].width = _column_width(df[column], column)
        lower = str(column).lower()
        if "revenue" in lower or "monetary value" in lower:
            number_format = "$#,##0.00"
        elif "share" in lower or "rate" in lower:
            number_format = "0.0%"
        elif column in {"Conversions", "Raw TTD rows", "Unique transaction IDs", "Matched transactions"}:
            number_format = "#,##0"
        else:
            number_format = None
        if number_format:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col_idx).number_format = number_format


def _column_width(series: pd.Series | pd.DataFrame, column_name: str) -> int:
    if isinstance(series, pd.DataFrame):
        sample_values = series.head(1000).to_numpy().ravel()
        max_value_width = max((len(str(value)) for value in sample_values), default=0)
    else:
        sample = series.head(1000)
        max_value_width = int(sample.map(lambda value: len(str(value))).max()) if not sample.empty else 0
    return min(max(max_value_width, len(str(column_name))) + 2, 48)


def _safe_sheet_name(sheet_name: str) -> str:
    invalid = "[]:*?/\\"
    safe = "".join("_" if char in invalid else char for char in sheet_name)
    return safe[:31]
